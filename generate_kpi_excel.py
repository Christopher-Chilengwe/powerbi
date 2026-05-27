from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

kb = [
    ("Marketing", [
        ("Brand reach", 5000, "Higher", "Total social media engagements + website visits per month", "Monthly"),
        ("Lead conversion rate", 0.1, "Higher", "Number of inquiries converted to applications ÷ total inquiries", "Monthly"),
        ("Event turnout", 0.75, "Higher", "Attendance at career fairs/open days ÷ target attendance", "Monthly"),
        ("Cost per lead", 50, "Lower", "Marketing spend ÷ # of new inquiries", "Monthly"),
    ]),
    ("Admissions", [
        ("Application-to-enrollment rate", 0.6, "Higher", "Enrolled applicants ÷ total applications", "Monthly"),
        ("Program fill rate", 0.9, "Higher", "Filled program slots ÷ total available slots", "Monthly"),
        ("Yield rate", 0.75, "Higher", "Accepted students who register ÷ total offers issued", "Monthly"),
        ("Processing time", 14, "Lower", "Average days from application submission to admission decision", "Monthly"),
    ]),
    ("Exams", [
        ("Exam completion rate", 0.98, "Higher", "Registered students who sit exams ÷ total registered", "Monthly"),
        ("Results turnaround time", 21, "Lower", "Days from exam administration to published results", "Monthly"),
        ("Malpractice rate", 0.005, "Lower", "Confirmed malpractice cases ÷ total exam candidates", "Monthly"),
        ("Appeal rate", 0.02, "Lower", "Exam appeals lodged ÷ students examined", "Monthly"),
    ]),
    ("Transportation", [
        ("Bus punctuality rate", 0.95, "Higher", "On-time trips ÷ total scheduled trips", "Monthly"),
        ("Fleet availability", 0.9, "Higher", "Operational vehicles ÷ total fleet", "Daily"),
        ("Cost per km per student", 0.5, "Lower", "Transport cost ÷ total student-km serviced", "Monthly"),
        ("Incident rate", 0.5, "Lower", "Accidents/breakdowns per month", "Monthly"),
    ]),
    ("Rock Farm", [
        ("Yield per hectare", 4, "Higher", "Agricultural output in tons per hectare", "Monthly"),
        ("Revenue vs cost", 1.2, "Higher", "Farm income ÷ farm operational expenses", "Quarterly"),
        ("Student participation", 120, "Higher", "Number of students engaged in farm activities", "Monthly"),
        ("Wastage rate", 0.05, "Lower", "Post-harvest losses ÷ total harvest", "Monthly"),
    ]),
    ("Skills - Short Courses", [
        ("Enrollment rate", 0.85, "Higher", "Seats filled ÷ course capacity", "Monthly"),
        ("Completion rate", 0.9, "Higher", "Students completing the course ÷ enrolled students", "Monthly"),
        ("Trainee satisfaction", 4.0, "Higher", "Average feedback rating out of 5", "Quarterly"),
        ("Employment uplift", 0.6, "Higher", "Trainees employed or upgraded within 3 months ÷ total trainees", "Quarterly"),
    ]),
    ("School of Health Sciences", [
        ("Clinical placement rate", 0.95, "Higher", "Students placed in clinical attachments ÷ cohort size", "Semesterly"),
        ("Licensure pass rate", 0.85, "Higher", "Graduates passing professional exams ÷ candidates", "Annual"),
        ("Patient contact hours", 180, "Higher", "Average clinical hours per student", "Semesterly"),
        ("Accreditation compliance", 1.0, "Higher", "Accreditation standards met ÷ total required standards", "Annual"),
    ]),
    ("Human Resource", [
        ("Staff turnover rate", 0.08, "Lower", "Staff departures ÷ average staff headcount", "Annual"),
        ("Vacancy fill time", 45, "Lower", "Average days to fill open positions", "Monthly"),
        ("Training completion", 0.9, "Higher", "Staff completing mandatory training ÷ total staff", "Annual"),
        ("Staff satisfaction score", 4.2, "Higher", "Average staff survey score out of 5", "Annual"),
    ]),
    ("Legal", [
        ("Case resolution time", 30, "Lower", "Average days to close legal matters", "Monthly"),
        ("Compliance rate", 1.0, "Higher", "Policies in compliance ÷ total reviewed policies", "Quarterly"),
        ("Contract review turnaround", 14, "Lower", "Average days to review and approve contractual documents", "Monthly"),
        ("Litigation cost", 0.02, "Lower", "Legal expenses ÷ total operating budget", "Annual"),
    ]),
    ("Library", [
        ("Utilization rate", 120, "Higher", "Average daily student visits", "Monthly"),
        ("E-resource usage", 800, "Higher", "Monthly downloads/logins of digital resources", "Monthly"),
        ("Book-to-student ratio", 0.5, "Higher", "Total available books ÷ enrolled students", "Annual"),
        ("Catalog update rate", 0.95, "Higher", "New titles cataloged within 30 days ÷ new acquisitions", "Quarterly"),
    ]),
    ("Accounts", [
        ("Fee collection rate", 0.95, "Higher", "Fees collected ÷ fees billed", "Monthly"),
        ("Days sales outstanding", 30, "Lower", "Average days to collect tuition and fees", "Monthly"),
        ("Budget variance", 0.05, "Lower", "Absolute difference between actual and budget ÷ budget", "Quarterly"),
        ("Audit findings", 0, "Lower", "Open audit issues requiring action", "Quarterly"),
    ]),
    ("School of Business and Humanities", [
        ("Program pass rate", 0.78, "Higher", "Students passing courses ÷ students registered", "Semesterly"),
        ("Research output", 20, "Higher", "Publications and conference contributions per year", "Annual"),
        ("Industry linkage", 5, "Higher", "Memoranda of Understanding signed with employers", "Annual"),
        ("Internship placement rate", 0.8, "Higher", "Students placed in industry internships ÷ eligible students", "Semesterly"),
    ]),
    ("School of Natural Science", [
        ("Lab usage rate", 0.85, "Higher", "Scheduled lab hours used ÷ total scheduled lab hours", "Monthly"),
        ("Research grants", 200000, "Higher", "Value of research grants secured", "Annual"),
        ("Project completion rate", 0.95, "Higher", "Final year projects completed on schedule ÷ total projects", "Annual"),
        ("Publication rate", 1.2, "Higher", "Peer-reviewed papers per lecturer per year", "Annual"),
    ]),
    ("School of Agriculture", [
        ("Field practical completion", 0.9, "Higher", "Students completing practical fieldwork ÷ total students", "Semesterly"),
        ("Crop/livestock yield", 3, "Higher", "Average output in tons per student plot", "Annual"),
        ("Extension outreach", 200, "Higher", "Farmers reached through outreach programs", "Annual"),
        ("Technology adoption rate", 0.8, "Higher", "Students using new agricultural methods ÷ total students", "Annual"),
    ]),
    ("School of Education", [
        ("Teaching practice completion", 0.92, "Higher", "Students passing teaching practice ÷ total participants", "Semesterly"),
        ("Graduate employment in schools", 0.7, "Higher", "Graduates employed in education roles within 1 year ÷ graduates", "Annual"),
        ("Compliance rate", 1.0, "Higher", "Programs meeting NQA standards ÷ total programs", "Annual"),
        ("Student evaluation score", 4.0, "Higher", "Average teaching evaluation score out of 5", "Semesterly"),
    ]),
    ("Quality Assurance/Auditors Office", [
        ("Program review cycle", 1.0, "Higher", "Programs reviewed on schedule ÷ total programs", "Annual"),
        ("Non-conformance rate", 2, "Lower", "QA issues per department identified during reviews", "Quarterly"),
        ("Action closure rate", 0.9, "Higher", "Audit findings resolved within 90 days ÷ total findings", "Quarterly"),
        ("Student satisfaction", 4.0, "Higher", "Average student satisfaction rating across reviews", "Annual"),
    ]),
    ("Projects", [
        ("Project completion rate", 0.9, "Higher", "Projects delivered by deadline ÷ total active projects", "Quarterly"),
        ("Donor disbursement rate", 0.95, "Higher", "Funds received ÷ funds pledged", "Quarterly"),
        ("Beneficiary reach", 5000, "Higher", "People impacted by project activities", "Quarterly"),
        ("Sustainability score", 0.8, "Higher", "Projects still operational 12 months after close ÷ total closed projects", "Annual"),
    ]),
    ("Sports", [
        ("Student participation", 0.15, "Higher", "Students engaged in sports ÷ total student population", "Monthly"),
        ("Inter-university wins", 5, "Higher", "Medals or trophies won per year", "Annual"),
        ("Injury rate", 0.015, "Lower", "Sports injuries ÷ athletes participating", "Monthly"),
        ("Facility usage", 120, "Higher", "Hours of sports facility use per week", "Weekly"),
    ]),
    ("Industrial Attachment", [
        ("Placement rate", 0.85, "Higher", "Students placed with approved employers ÷ eligible students", "Semesterly"),
        ("Supervisor rating", 4.0, "Higher", "Average employer supervision score out of 5", "Semesterly"),
        ("Conversion to employment", 0.4, "Higher", "Students offered jobs after attachment ÷ attached students", "Annual"),
        ("Attachment completion", 0.98, "Higher", "Students completing the full attachment period ÷ all attached students", "Semesterly"),
    ]),
    ("Rockmas Trust Schools", [
        ("Learner pass rate", 0.85, "Higher", "Learners passing national exams ÷ learners sitting exams", "Annual"),
        ("Teacher-pupil ratio", 30, "Lower", "Average number of pupils per teacher", "Annual"),
        ("Retention rate", 0.95, "Higher", "Pupils progressing to next grade ÷ enrolled pupils", "Annual"),
        ("Fee collection rate", 0.98, "Higher", "School fees collected ÷ billed fees", "Monthly"),
    ]),
    ("Research & Innovation", [
        ("Research funding growth", 0.15, "Higher", "Year-on-year increase in research funding", "Annual"),
        ("Patent/prototype count", 6, "Higher", "New patents or prototypes developed", "Annual"),
        ("Collaborative projects", 8, "Higher", "Joint research or innovation partnerships active", "Annual"),
        ("Innovation training completion", 0.9, "Higher", "Participants completing innovation workshops ÷ enrolled participants", "Annual"),
    ]),
]

wb = Workbook()
ws = wb.active
ws.title = "Data Entry"
headers = ["Department","Indicator","Target","Goal","Definition/Formula","Frequency","Responsible","Jan","Feb","Mar","Q1 Avg","Status","Notes"]
ws.append(headers)
for cell in ws[1]:
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='center', vertical='center')

row = 2
for dept, entries in kb:
    for indicator, target, goal, definition, freq in entries:
        ws.append([dept, indicator, target, goal, definition, freq, "", None, None, None, None, None, ""])
        row += 1

for col, width in [(1,20),(2,35),(3,14),(4,12),(5,50),(6,12),(7,18),(8,12),(9,12),(10,12),(11,12),(12,14),(13,20)]:
    ws.column_dimensions[get_column_letter(col)].width = width

for r in range(2, row):
    ws[f'K{r}'] = f"=AVERAGE(H{r}:J{r})"
    formula = "=IF(OR(K{0}='',C{0}=''),'',IF(D{0}='Higher',IF(K{0}>=C{0},'On Target','Below Target'),IF(K{0}<=C{0},'On Target','Below Target')))".format(r)
    ws[f'L{r}'] = formula

for r in range(2, row):
    cell = ws.cell(row=r, column=3)
    if isinstance(cell.value, float) and cell.value <= 1:
        cell.number_format = '0.00%'

sd = wb.create_sheet(title="Summary Dashboard")
summary_headers = ["Department","KPI Count","On Target","Below Target","% On Target","Notes"]
sd.append(summary_headers)
for cell in sd[1]:
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='center', vertical='center')

for idx, (dept, _) in enumerate(kb, start=2):
    sd[f'A{idx}'] = dept
    sd[f'B{idx}'] = f"=COUNTIF('Data Entry'!A:A,A{idx})"
    sd[f'C{idx}'] = f"=COUNTIFS('Data Entry'!A:A,A{idx},'Data Entry'!L:L,\"On Target\")"
    sd[f'D{idx}'] = f"=COUNTIFS('Data Entry'!A:A,A{idx},'Data Entry'!L:L,\"Below Target\")"
    sd[f'E{idx}'] = f"=IF(B{idx}=0,'',C{idx}/B{idx})"

for col, width in [(1,28),(2,12),(3,12),(4,12),(5,14),(6,28)]:
    sd.column_dimensions[get_column_letter(col)].width = width
for idx in range(2, 2+len(kb)):
    sd[f'E{idx}'].number_format = '0.00%'

wb.save('Rockview_University_KPI_Template.xlsx')
print('Created Rockview_University_KPI_Template.xlsx')
